import os
from openpyxl import load_workbook, Workbook

folder_path = "vystupni_data"
upravene_soubory = []

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx") and not filename.startswith("upraveno_") and not filename.startswith("vse_komplet"):
        file_path = os.path.join(folder_path, filename)
        print(f"Zpracovávám soubor: {filename}")

        workbook = load_workbook(file_path)

        if "Komplet" in workbook.sheetnames:
            del workbook["Komplet"]

        komplet_sheet = workbook.create_sheet("Komplet")

        ranges_to_copy = {
            "Table_3": (21, 37),
            "Table_4": (7, 10),
            "Table_6": (27, 35),
            "Table_7": (7, 29),
            "Table_8": (7, 35),
            "Table_9": (7, 20),
            "Table_10": (7, 37),
            # "Table_11": (7, 15),
        }

        for sheet_name, (start_row, end_row) in ranges_to_copy.items():
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                    komplet_sheet.append(row)

        for i in range(1, 22):
            name = f"Table_{i}"
            if name in workbook.sheetnames:
                del workbook[name]

        upraveny_nazev = f"upraveno_{filename}"
        upraveny_path = os.path.join(folder_path, upraveny_nazev)
        workbook.save(upraveny_path)

        upravene_soubory.append(upraveny_path)

print("Spojuji listy 'Komplet' do jednoho souboru...")

import re

def natural_key(string):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', string)]

upravene_soubory.sort(key=lambda x: natural_key(os.path.basename(x)))


vystupni_workbook = Workbook()
vystupni_list = vystupni_workbook.active
vystupni_list.title = "VseKomplet"

start_col = 1

for file_path in upravene_soubory:
    wb = load_workbook(file_path, data_only=True)
    if "Komplet" in wb.sheetnames:
        ws = wb["Komplet"]
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_index, cell_value in enumerate(row):
                vystupni_list.cell(row=row_index, column=start_col + col_index, value=cell_value)
        start_col += ws.max_column

vystupni_soubor = os.path.join(folder_path, "vse_komplet.xlsx")
vystupni_workbook.save(vystupni_soubor)

print(f"Hotovo! Všechny soubory byly zpracovány a sloučeny do: {vystupni_soubor}")
